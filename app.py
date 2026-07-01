import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta
from io import BytesIO
import math

st.set_page_config(page_title="선택적 근무 · 초과시간 관리", page_icon="⏱", layout="wide")
st.markdown("""
<style>
[data-testid="stFileUploaderDropzoneInstructions"] { display: none; }
[data-testid="stFileUploaderDropzone"] > div:last-child { display: none; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');
html, body, [class*="css"] { font-family: 'Noto Sans KR', sans-serif; }
.app-header {
    background: linear-gradient(135deg, #1F3864 0%, #2F5496 100%);
    border-radius: 12px; padding: 28px 32px; margin-bottom: 28px; color: white;
}
.app-header h1 { font-size: 22px; font-weight: 700; margin: 0 0 6px 0; }
.app-header p  { font-size: 13px; margin: 0; opacity: 0.75; }
.rule-box {
    background: #F0F4FB; border-left: 4px solid #2F5496;
    border-radius: 0 8px 8px 0; padding: 14px 18px;
    font-size: 13px; color: #2C3E50; line-height: 1.8; margin-bottom: 24px;
}
.card-wrap { display: flex; gap: 14px; flex-wrap: wrap; margin-bottom: 24px; }
.summary-card {
    background: white; border: 1px solid #E0E8F5; border-radius: 10px;
    padding: 16px 22px; min-width: 160px; flex: 1;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
.summary-card .label { font-size: 11px; color: #7F8C8D; font-weight: 500; margin-bottom: 6px; }
.summary-card .value { font-size: 22px; font-weight: 700; color: #1F3864; }
.summary-card .sub   { font-size: 11px; color: #95A5A6; margin-top: 2px; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="app-header">
    <h1>⏱ 선택적 근무시간제 · 초과시간 관리</h1>
    <p>인트라넷 근태현황 파일을 업로드하면 인정근무시간과 잔여시간을 자동 계산합니다</p>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="rule-box">
    📌 <strong>계산 기준</strong> &nbsp;|&nbsp;
    인정출근 = 실제 출근시각 <strong>30분 올림</strong> (08:00 이전은 08:00 고정) &nbsp;·&nbsp;
    인정퇴근 = 실제 퇴근시각 <strong>30분 내림</strong> &nbsp;·&nbsp;
    인정근무 = 체류시간 − 점심 1H &nbsp;·&nbsp;
    8H 초과 → <strong>+적립</strong> / 8H 미달 → <strong>−차감</strong> (30분 단위)
</div>
""", unsafe_allow_html=True)

# ── 헬퍼 함수 ──────────────────────────────────────────────────
def parse_hms(s):
    if pd.isna(s) or str(s).strip() in ('', 'False', 'NaT'):
        return None
    try:
        p = str(s).strip().split(':')
        return timedelta(hours=int(p[0]), minutes=int(p[1]), seconds=int(p[2]))
    except:
        return None

def ceil30(td):
    m = int(td.total_seconds() // 60)
    return timedelta(minutes=math.ceil(m / 30) * 30)

def floor30(td):
    m = int(td.total_seconds() // 60)
    return timedelta(minutes=(m // 30) * 30)

def fmt_td(td):
    if td is None: return ''
    m = int(td.total_seconds() // 60)
    return f"{m // 60}:{m % 60:02d}"

def fmt_net(minutes):
    h = abs(minutes) // 60
    m = abs(minutes) % 60
    if minutes > 0:   return f"+{h}:{m:02d}"
    elif minutes < 0: return f"-{h}:{m:02d}"
    else:             return "0:00"

EIGHT_AM = timedelta(hours=8)
LUNCH    = timedelta(hours=1)
EIGHT_H  = timedelta(hours=8)

# ── 파일 타입 감지 ─────────────────────────────────────────────
def detect_file_type(xl):
    """부서 파일 vs 개인 파일 구분 - P_ 시트 우선"""
    sheets = xl.sheet_names
    # P_ 시트 우선 탐색
    for sheet in sheets:
        if sheet.startswith('P_') and '근태현황' in sheet:
            return 'team', sheet
    for sheet in sheets:
        if '기간별 근태현황' in sheet:
            return 'team', sheet
        if '월간 근태현황' in sheet or ('근태현황' in sheet and not sheet.startswith('U_')):
            return 'personal', sheet
    return 'unknown', sheets[0]

# ── 계산 로직 (공통) ───────────────────────────────────────────
FOUR_H = timedelta(hours=4)

def calc_row(in_r, out_r, vac):
    # 반차 여부: 총휴가시간이 4H인 경우
    is_half = vac is not None and abs(vac.total_seconds() - 4*3600) < 60

    if in_r is None or out_r is None:
        if is_half:
            note = '반차'
        elif vac and vac.total_seconds() > 0:
            note = '휴무/휴가'
        else:
            note = '휴무'
        return None, None, None, None, 0, note

    adj_in  = EIGHT_AM if in_r < EIGHT_AM else ceil30(in_r)
    adj_out = floor30(out_r)
    stay    = max(adj_out - adj_in, timedelta(0))

    if is_half:
        # 반차: 체류 4H 이하 → 점심 미공제 / 4H 초과 → 점심 1H 공제, 기준 4H
        work   = stay if stay <= FOUR_H else stay - LUNCH
        base_h = FOUR_H
        note   = '반차'
    else:
        work   = max(stay - LUNCH, timedelta(0))
        base_h = EIGHT_H
        note   = '08:00 이전→보정' if in_r < EIGHT_AM else ''

    diff_min = int((work - base_h).total_seconds() // 60)
    net_min  = (abs(diff_min) // 30) * 30
    net_min  = net_min if diff_min >= 0 else -net_min
    return adj_in, adj_out, work, fmt_net(net_min), net_min, note

# ── 팀 파일 파싱 ───────────────────────────────────────────────
def parse_team(df_all):
    try:    period = str(df_all.iloc[4, 2])
    except: period = ''

    # 헤더 행 동적 탐색
    header_row = 6  # fallback
    for i in range(10):
        if '일자' in df_all.iloc[i].tolist():
            header_row = i
            break

    headers = df_all.iloc[header_row].tolist()

    def find_col(candidates):
        for name in candidates:
            for i, h in enumerate(headers):
                if isinstance(h, str) and name in h:
                    return i
        return None

    col_date = find_col(['일자'])
    col_name = find_col(['이름'])
    col_rank = find_col(['직위'])
    col_in   = find_col(['출근시간'])
    col_out  = find_col(['퇴근시간'])
    col_vac  = find_col(['총 휴가시간', '총휴가시간'])

    df_data = df_all.iloc[header_row+1:, [col_date, col_name, col_rank, col_in, col_out, col_vac]].copy()
    df_data.columns = ['일자', '이름', '직위', '출근시간', '퇴근시간', '총휴가시간']
    df_data = df_data[df_data['일자'].notna()].reset_index(drop=True)
    return df_data, period

# ── 개인 파일 파싱 ─────────────────────────────────────────────
def parse_personal(df_all, sheet_name):
    try:    period = str(df_all.iloc[4, 2])
    except: period = ''
    # 이름 추출 (시트명: "김지윤 월간 근태현황")
    name = sheet_name.replace('월간 근태현황', '').replace('근태현황', '').strip()
    # B=일자, C=출근, F=퇴근, P=휴가 (0-indexed: 1,2,5,15)
    df_data = df_all.iloc[7:, [1, 2, 5, 15]].copy()
    df_data.columns = ['일자', '출근시간', '퇴근시간', '총휴가시간']
    df_data['이름'] = name
    df_data['직위'] = ''
    df_data = df_data[df_data['일자'].notna()].reset_index(drop=True)
    return df_data, period, name

# ── 공통 계산 ──────────────────────────────────────────────────
def process(df_raw):
    detail_rows = []
    for _, r in df_raw.iterrows():
        date_str = str(r['일자'])
        emp_name = str(r['이름'])
        emp_rank = str(r.get('직위', ''))
        in_r     = parse_hms(r['출근시간'])
        out_r    = parse_hms(r['퇴근시간'])
        vac      = parse_hms(r['총휴가시간'])

        if not emp_name or emp_name in ('nan', '이름', ''):
            continue

        adj_in, adj_out, work, net_str, net_min, note = calc_row(in_r, out_r, vac)

        detail_rows.append({
            '일자': date_str, '이름': emp_name, '직위': emp_rank,
            '실출근': str(r['출근시간']) if in_r else '',
            '실퇴근': str(r['퇴근시간']) if out_r else '',
            '인정출근': fmt_td(adj_in), '인정퇴근': fmt_td(adj_out),
            '인정근무시간': fmt_td(work),
            '8H 대비': net_str if net_str else '',
            'net_min': net_min, '비고': note
        })

    detail = pd.DataFrame(detail_rows)

    summary_rows = []
    for name in detail['이름'].unique():
        g = detail[detail['이름'] == name]
        rank_vals = g[g['직위'].notna() & (g['직위'] != 'nan') & (g['직위'] != '')]['직위']
        rank      = rank_vals.iloc[0] if len(rank_vals) > 0 else ''
        total_min = int(g['net_min'].sum())
        blocks    = total_min // 30
        if blocks > 0:   bigo = f"{blocks}회 × 30분 조기퇴근 사용 가능"
        elif blocks < 0: bigo = f"{abs(blocks)}회 × 30분 추가 근무 필요"
        else:            bigo = '-'
        summary_rows.append({
            '이름': name, '직위': rank,
            '누적 잔여시간': fmt_net(total_min),
            '사용가능 블록(30분)': blocks,
            '비고': bigo
        })

    return detail, pd.DataFrame(summary_rows)

# ── Excel 출력 ─────────────────────────────────────────────────
def to_excel(detail, summary):
    wb  = Workbook()
    C_BLU = "2F5496"
    C_YEL = "FFF2CC"; C_PNK = "FCE4EC"; C_GRN = "E2EFDA"; C_BRD = "B8CCE4"
    thin = Side(style='thin', color=C_BRD)
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hc(ws, row, col, val, bg="1F3864", fg="FFFFFF", sz=10, bold=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Arial', bold=bold, color=fg, size=sz)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = brd
        return c

    def dc(ws, row, col, val, bg=None, bold=False, align='center', sz=9, fg="000000"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Arial', bold=bold, size=sz, color=fg)
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border    = brd
        if bg: c.fill = PatternFill('solid', fgColor=bg)
        return c

    # Sheet1: 누적잔여 요약
    ws1 = wb.active; ws1.title = "누적잔여 요약"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 8
    ws1.column_dimensions['A'].width = 3
    ws1.merge_cells('B2:G2')
    c = ws1['B2']; c.value = "선택적 근무시간제 · 잔여시간 현황"
    c.font = Font(name='Arial', bold=True, size=14, color=C_BLU)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[2].height = 28
    ws1.merge_cells('B3:G3')
    c = ws1['B3']; c.value = "※ 인정출근 30분 올림(08:00 이전→08:00) | 인정퇴근 30분 내림 | 점심 1H 공제 | 8H 초과→+적립 / 미달→−차감 (30분 단위)"
    c.font = Font(name='Arial', size=8, color="595959")
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[3].height = 15; ws1.row_dimensions[4].height = 6; ws1.row_dimensions[5].height = 32

    for col, h, w in zip([2,3,4,5,6], ['이름','직위','누적 잔여시간','사용가능 블록(30분)','비고'], [14,10,16,20,34]):
        hc(ws1, 5, col, h); ws1.column_dimensions[get_column_letter(col)].width = w

    for i, row in summary.iterrows():
        r = i + 6; ws1.row_dimensions[r].height = 24
        bg  = C_GRN if i % 2 == 0 else None
        blk = row['사용가능 블록(30분)']
        dc(ws1, r, 2, row['이름'],  bg=bg, bold=True, sz=10)
        dc(ws1, r, 3, row['직위'],  bg=bg, sz=9)
        if blk > 0:
            dc(ws1, r, 4, row['누적 잔여시간'], bg=C_YEL, bold=True, sz=11, fg="7F4F00")
            dc(ws1, r, 5, blk, bg=C_YEL, bold=True, sz=10, fg="7F4F00")
        elif blk < 0:
            dc(ws1, r, 4, row['누적 잔여시간'], bg=C_PNK, bold=True, sz=11, fg="8B0000")
            dc(ws1, r, 5, blk, bg=C_PNK, bold=True, sz=10, fg="8B0000")
        else:
            dc(ws1, r, 4, row['누적 잔여시간'], bg=bg, sz=10)
            dc(ws1, r, 5, blk, bg=bg, sz=10)
        dc(ws1, r, 6, row['비고'], bg=bg, sz=9, align='left')
    ws1.freeze_panes = 'B6'

    # Sheet2: 일별 상세
    ws2 = wb.create_sheet("일별 상세")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 8
    ws2.column_dimensions['A'].width = 3; ws2.column_dimensions['L'].width = 3
    ws2.merge_cells('B2:K2')
    c = ws2['B2']; c.value = "선택적 근무시간제 · 일별 인정근무시간 상세"
    c.font = Font(name='Arial', bold=True, size=13, color=C_BLU)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[2].height = 26
    ws2.merge_cells('B3:K3')
    c = ws2['B3']; c.value = "※ 인정근무 = (인정퇴근−인정출근)−점심1H | +초과(노랑) / −미달(분홍) 표시 | 30분 단위 절사"
    c.font = Font(name='Arial', size=8, color="595959")
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[3].height = 15; ws2.row_dimensions[4].height = 6; ws2.row_dimensions[5].height = 32

    hdrs = ['일자','이름','직위','실출근','실퇴근','인정출근','인정퇴근','인정근무시간','8H 대비','비고']
    wids = [12,12,8,11,11,11,11,14,14,18]
    for col, h, w in zip(range(2,12), hdrs, wids):
        hc(ws2, 5, col, h); ws2.column_dimensions[get_column_letter(col)].width = w

    color_pool = ["EBF3FB","FFF9F0","F0FBF0","FDF0FB","FFF5E6","F5F0FF"]
    person_colors = {name: color_pool[i % len(color_pool)] for i, name in enumerate(detail['이름'].unique())}

    for i, row in detail.iterrows():
        r = i + 6; ws2.row_dimensions[r].height = 18
        base_bg = person_colors.get(row['이름'], "FFFFFF")
        net     = row['net_min']
        vals    = [row['일자'], row['이름'], row['직위'], row['실출근'], row['실퇴근'],
                   row['인정출근'], row['인정퇴근'], row['인정근무시간'], row['8H 대비'], row['비고']]
        for col_idx, val in enumerate(vals):
            col = col_idx + 2
            if col in [9,10] and net > 0:
                dc(ws2, r, col, val, bg=C_YEL, bold=True, sz=9, fg="7F4F00")
            elif col in [9,10] and net < 0:
                dc(ws2, r, col, val, bg=C_PNK, bold=True, sz=9, fg="8B0000")
            elif col == 11 and '보정' in str(row['비고']):
                dc(ws2, r, col, val, bg=C_PNK, sz=8, align='left')
            elif col in [5,6]:
                dc(ws2, r, col, val, bg="F5F5F5", sz=8)
            else:
                dc(ws2, r, col, val, bg=base_bg, sz=9)
    ws2.freeze_panes = 'B6'

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ── Session state ─────────────────────────────────────────────
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
if 'remove_file' not in st.session_state:
    st.session_state.remove_file = False
if 'uploader_key' not in st.session_state:
    st.session_state.uploader_key = 0

if st.session_state.remove_file:
    st.session_state.uploaded_file = None
    st.session_state.remove_file = False
    st.session_state.uploader_key += 1
    st.rerun()

# ── 사이드바 ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown('''
<p style="font-size:11px;font-weight:600;color:#7F8C8D;letter-spacing:0.6px;text-transform:uppercase;margin:0 0 10px;">파일 업로드</p>
<p style="font-size:12px;color:#444444;line-height:1.7;margin:0 0 12px;">
    📂 인트라넷에서 다운받은 RAW 엑셀데이터를<br>파일명 변경, 가공없이 그대로 업로드해주세요.
</p>
''', unsafe_allow_html=True)

    new_file = st.file_uploader(
        "개인 월간 파일 또는 팀 전체 파일(이름 마스킹처리 권장)을 업로드하세요.",
        type=['xlsx'],
        key=f"file_uploader_{st.session_state.uploader_key}"
    )
    if new_file is not None:
        st.session_state.uploaded_file = new_file

    if st.session_state.uploaded_file is not None:
        st.markdown(f'''
<div style="background:#EBF3FB;border-radius:8px;padding:8px 12px;margin:8px 0;display:flex;align-items:center;gap:8px;">
    <span style="font-size:16px;">📄</span>
    <span style="font-size:12px;color:#1F3864;font-weight:500;flex:1;word-break:break-all;">{st.session_state.uploaded_file.name}</span>
</div>
''', unsafe_allow_html=True)
        remove = st.button("✕  파일 제거", use_container_width=True)
    else:
        remove = False

    st.markdown('''
<hr style="margin:20px 0 12px;border:none;border-top:0.5px solid #E0E8F5;">
<p style="font-size:11px;color:#95A5A6;line-height:1.8;margin:0;">
    <strong style="color:#7F8C8D;">계산 기준</strong><br>
    출근 30분 올림<br>(08:00 이전 → 08:00 고정)<br>
    퇴근 30분 내림<br>
    점심 1H 공제<br>
    8H 초과 → +적립<br>
    8H 미달 → −차감 (30분 단위)
</p>
<hr style="margin:12px 0;border:none;border-top:0.5px solid #E0E8F5;">
<p style="font-size:11px;color:#95A5A6;line-height:1.8;margin:0;">
    <strong style="color:#7F8C8D;">반차 계산 기준</strong><br>
    기준근무: 4H<br>
    체류 4H 이하 → 점심 미공제<br>
    체류 4H 초과 → 점심 1H 공제<br><br>
    <span style="color:#B0B0B0;">예시</span><br>
    08:00 출근 → 12:00 퇴근<br>
    체류 4H → 인정 4H ✓<br><br>
    09:00 출근 → 14:00 퇴근<br>
    체류 5H − 1H → 인정 4H ✓
</p>
''', unsafe_allow_html=True)

if remove:
    st.session_state.remove_file = True
    st.rerun()

uploaded = st.session_state.uploaded_file

# ── 메인 UI ────────────────────────────────────────────────────
if uploaded:
    with st.spinner("계산 중..."):
        try:
            xl      = pd.ExcelFile(uploaded)
            f_type, sheet = detect_file_type(xl)
            df_all  = pd.read_excel(uploaded, sheet_name=sheet, header=None)

            if f_type == 'team':
                df_data, period = parse_team(df_all)
                mode_label = "팀 전체"
            elif f_type == 'personal':
                df_data, period, person_name = parse_personal(df_all, sheet)
                mode_label = f"개인 ({person_name})"
            else:
                st.error("파일 형식을 인식할 수 없습니다. 인트라넷 원본 파일인지 확인해주세요.")
                st.stop()

            detail, summary = process(df_data)

            # 지표
            total_people        = len(summary)
            people_with_surplus = len(summary[summary['사용가능 블록(30분)'] > 0])
            people_with_deficit = len(summary[summary['사용가능 블록(30분)'] < 0])
            best_idx            = summary['사용가능 블록(30분)'].values.argmax()
            max_val             = summary['누적 잔여시간'].iloc[best_idx]
            max_name            = summary['이름'].iloc[best_idx]

            # 모드 배지
            badge_color = "#1F6B2F" if f_type == 'team' else "#2F5496"
            st.markdown(f'<span style="background:{badge_color};color:white;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;">{"📋 팀 전체 모드" if f_type == "team" else "👤 개인 모드"}</span>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            st.markdown(f"""
            <div class="card-wrap">
                <div class="summary-card">
                    <div class="label">조회기간</div>
                    <div class="value" style="font-size:15px;">{period}</div>
                </div>
                <div class="summary-card">
                    <div class="label">{"대상 인원" if f_type == "team" else "이름"}</div>
                    <div class="value">{"{}명".format(total_people) if f_type == "team" else max_name}</div>
                </div>
                <div class="summary-card">
                    <div class="label">{"잔여시간 보유" if f_type == "team" else "누적 잔여시간"}</div>
                    <div class="value" style="color:#1F6B2F;">{"{}명".format(people_with_surplus) if f_type == "team" else max_val}</div>
                    <div class="sub">{"초과 적립자" if f_type == "team" else "사용가능 블록: {}회".format(summary["사용가능 블록(30분)"].iloc[best_idx])}</div>
                </div>
                {"" if f_type == "personal" else f'<div class="summary-card"><div class="label">차감 발생</div><div class="value" style="color:#C0392B;">{people_with_deficit}명</div><div class="sub">8H 미달 차감자</div></div>'}
                {"" if f_type == "personal" else f'<div class="summary-card"><div class="label">최대 누적</div><div class="value">{max_val}</div><div class="sub">{max_name}</div></div>'}
            </div>
            """, unsafe_allow_html=True)

            tab1, tab2, tab3 = st.tabs(["📊 누적잔여 요약", "📋 일별 상세", "📈 통계"])

            with tab1:
                st.dataframe(summary, use_container_width=True, hide_index=True,
                    column_config={
                        "이름":             st.column_config.TextColumn("이름", width=100),
                        "직위":             st.column_config.TextColumn("직위", width=80),
                        "누적 잔여시간":    st.column_config.TextColumn("누적 잔여시간", width=120),
                        "사용가능 블록(30분)": st.column_config.NumberColumn("사용가능 블록(30분)", width=160),
                        "비고":             st.column_config.TextColumn("비고", width=280),
                    })

            with tab2:
                if f_type == 'team':
                    col1, _ = st.columns([2, 5])
                    with col1:
                        names    = ['전체'] + list(detail['이름'].unique())
                        selected = st.selectbox("팀원 선택", names)
                    df_show = detail if selected == '전체' else detail[detail['이름'] == selected]
                else:
                    df_show = detail

                df_show = df_show.drop(columns=['net_min'], errors='ignore')
                st.dataframe(df_show, use_container_width=True, hide_index=True,
                    column_config={
                        "일자":         st.column_config.TextColumn("일자", width=100),
                        "이름":         st.column_config.TextColumn("이름", width=90),
                        "직위":         st.column_config.TextColumn("직위", width=70),
                        "실출근":       st.column_config.TextColumn("실출근", width=90),
                        "실퇴근":       st.column_config.TextColumn("실퇴근", width=90),
                        "인정출근":     st.column_config.TextColumn("인정출근", width=90),
                        "인정퇴근":     st.column_config.TextColumn("인정퇴근", width=90),
                        "인정근무시간": st.column_config.TextColumn("인정근무시간", width=110),
                        "8H 대비":      st.column_config.TextColumn("8H 대비(+초과/−미달)", width=150),
                        "비고":         st.column_config.TextColumn("비고", width=140),
                    })

            with tab3:
                if f_type == 'personal':
                    st.info("통계는 팀 전체 파일 업로드 시 제공됩니다.")
                else:
                    work_rows = detail[detail['인정근무시간'] != ''].copy()

                    def hm_to_min(s):
                        if not s or s == '': return None
                        try:
                            h, m = s.split(':')
                            return int(h)*60 + int(m)
                        except: return None

                    def min_to_hm(m):
                        if m is None: return '-'
                        return f"{int(m)//60}:{int(m)%60:02d}"

                    work_rows['work_min'] = work_rows['인정근무시간'].apply(hm_to_min)
                    work_rows = work_rows[work_rows['work_min'].notna()]
                    work_rows_full = work_rows[~work_rows['비고'].str.contains('휴가|반차', na=False)]

                    team_avg = work_rows['work_min'].mean()
                    team_max = work_rows['work_min'].max()
                    team_max_row = work_rows.loc[work_rows['work_min'].idxmax()]
                    team_min = work_rows_full['work_min'].min() if len(work_rows_full) > 0 else None
                    team_min_row = work_rows_full.loc[work_rows_full['work_min'].idxmin()] if len(work_rows_full) > 0 else None
                    team_surplus_days = int((work_rows['net_min'] > 0).sum())
                    team_deficit_days = int((work_rows['net_min'] < 0).sum())
                    min_name = team_min_row['이름'] if team_min_row is not None else '-'
                    min_date = team_min_row['일자'] if team_min_row is not None else ''

                    st.markdown(f'''
<div class="card-wrap">
    <div class="summary-card"><div class="label">팀 일 평균</div><div class="value">{min_to_hm(team_avg)}</div><div class="sub">조회기간 전체</div></div>
    <div class="summary-card"><div class="label">최장 근무일</div><div class="value" style="color:#2a78d6;">{min_to_hm(team_max)}</div><div class="sub">{team_max_row["이름"]} · {team_max_row["일자"]}</div></div>
    <div class="summary-card"><div class="label">최단 근무일</div><div class="value" style="color:#888;">{min_to_hm(team_min)}</div><div class="sub">{min_name} · {min_date}</div></div>
    <div class="summary-card"><div class="label">초과 적립일</div><div class="value" style="color:#1baf7a;">{team_surplus_days}일</div><div class="sub">팀 합산</div></div>
    <div class="summary-card"><div class="label">차감 발생일</div><div class="value" style="color:#e34948;">{team_deficit_days}일</div><div class="sub">팀 합산</div></div>
</div>
''', unsafe_allow_html=True)

                    colors = ['#2a78d6','#1baf7a','#eda100','#4a3aa7','#e34948','#eb6834']
                    person_stats = []
                    max_avg = 0
                    for name in detail['이름'].unique():
                        g = work_rows[work_rows['이름'] == name]
                        g_full = work_rows_full[work_rows_full['이름'] == name]
                        if len(g) == 0: continue
                        avg = g['work_min'].mean()
                        max_avg = max(max_avg, avg)
                        person_stats.append({
                            'name': name, 'avg': avg,
                            'max': g['work_min'].max(),
                            'min': g_full['work_min'].min() if len(g_full) > 0 else None,
                            'surplus_days': int((g['net_min'] > 0).sum()),
                            'deficit_days': int((g['net_min'] < 0).sum()),
                            'work_days': len(g)
                        })

                    st.markdown('<div style="background:#f8f9fa;border-radius:12px;padding:1.25rem;margin-bottom:16px;">', unsafe_allow_html=True)
                    st.markdown('<div style="font-size:13px;font-weight:500;margin-bottom:16px;">팀원별 일 평균 근무시간</div>', unsafe_allow_html=True)
                    for i, ps in enumerate(person_stats):
                        bar_pct = int((ps['avg'] / max(max_avg, 600)) * 100)
                        base_pct = int((480 / max(max_avg, 600)) * 100)
                        color = colors[i % len(colors)]
                        st.markdown(f'''
<div style="margin-bottom:16px;">
    <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:6px;">
        <span style="color:#555;">{ps["name"]}</span><span style="font-weight:500;">{min_to_hm(ps["avg"])}</span>
    </div>
    <div style="position:relative;background:#e0e0e0;border-radius:4px;height:10px;">
        <div style="width:{bar_pct}%;height:100%;background:{color};border-radius:4px;"></div>
        <div style="position:absolute;left:{base_pct}%;top:-3px;width:1px;height:16px;background:#999;"></div>
    </div>
    <div style="display:flex;justify-content:flex-end;gap:12px;margin-top:4px;font-size:11px;color:#888;">
        <span>초과 <span style="color:#1baf7a;font-weight:500;">{ps["surplus_days"]}일</span></span>
        <span>차감 <span style="color:#e34948;font-weight:500;">{ps["deficit_days"]}일</span></span>
    </div>
</div>
''', unsafe_allow_html=True)

                    st.markdown('''
<div style="padding-top:10px;border-top:0.5px solid #ddd;">
    <div style="display:flex;align-items:center;gap:6px;margin-bottom:8px;">
        <div style="width:1px;height:12px;background:#999;"></div>
        <span style="font-size:11px;color:#888;">8H 기준선</span>
    </div>
    <div style="font-size:11px;color:#aaa;line-height:1.8;">
        일 평균 · 최장 근무: 반차 포함하여 계산<br>
        최단 근무: 연차 · 반차 불포함하여 계산<br>
        초과일 · 차감일: 인정근무 기준 8H 대비 30분 단위 가감일 수
    </div>
</div></div>
''', unsafe_allow_html=True)

                    stats_df = pd.DataFrame([{
                        '이름': ps['name'], '일 평균': min_to_hm(ps['avg']),
                        '최장 근무': min_to_hm(ps['max']), '최단 근무': min_to_hm(ps['min']),
                        '초과일': ps['surplus_days'], '차감일': ps['deficit_days'], '근무일수': ps['work_days']
                    } for ps in person_stats])
                    st.dataframe(stats_df, use_container_width=True, hide_index=True,
                        column_config={
                            '이름': st.column_config.TextColumn('이름', width=100),
                            '일 평균': st.column_config.TextColumn('일 평균', width=90),
                            '최장 근무': st.column_config.TextColumn('최장 근무', width=90),
                            '최단 근무': st.column_config.TextColumn('최단 근무', width=90),
                            '초과일': st.column_config.NumberColumn('초과일', width=70),
                            '차감일': st.column_config.NumberColumn('차감일', width=70),
                            '근무일수': st.column_config.NumberColumn('근무일수', width=80),
                        })

            st.divider()
            fname     = f"선택적근무_잔여시간_{period.replace(' ','').replace('~','_')}.xlsx" if period else "선택적근무_잔여시간.xlsx"
            excel_buf = to_excel(detail, summary)
            st.download_button(label="⬇️ Excel 다운로드", data=excel_buf, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.error(f"파일 처리 중 오류가 발생했습니다: {e}")
            st.info("인트라넷 원본 파일인지 확인해주세요.")
else:
    st.info("⬆️ 부서 전체 파일 또는 개인 월간 파일을 업로드하세요.")
